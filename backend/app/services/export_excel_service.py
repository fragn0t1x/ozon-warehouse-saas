from __future__ import annotations

import os
from calendar import month_name
from datetime import datetime, timezone
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.export_status import ensure_export_root_dir


THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)
HEADER_FILL = PatternFill("solid", fgColor="E0F2FE")
SUBHEADER_FILL = PatternFill("solid", fgColor="ECFDF3")
SECTION_FILL = PatternFill("solid", fgColor="F8FAFC")
TITLE_FONT = Font(bold=True, size=13)
HEADER_FONT = Font(bold=True, size=11)
CELL_ALIGN = Alignment(vertical="center", horizontal="center", wrap_text=True)
TEXT_ALIGN = Alignment(vertical="center", horizontal="left", wrap_text=True)


def _safe_sheet_name(name: str, fallback: str) -> str:
    normalized = str(name or "").strip() or fallback
    forbidden = ['\\', '/', '*', '?', ':', '[', ']']
    for char in forbidden:
        normalized = normalized.replace(char, " ")
    normalized = normalized[:31].strip()
    return normalized or fallback


def _write_row(sheet, row_idx: int, values: Iterable, *, fill=None, font=None, alignment=None) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = alignment or CELL_ALIGN
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font


def _set_widths(sheet, widths: dict[int, float]) -> None:
    for col_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _display_color(color: str) -> str:
    normalized = str(color or "").strip()
    return normalized if normalized and normalized.lower() != "без цвета" else ""


def _variant_color(attributes: dict[str, Any] | None) -> str:
    return str((attributes or {}).get("Цвет") or "Без цвета")


def _variant_size(attributes: dict[str, Any] | None) -> str:
    return str((attributes or {}).get("Размер") or "Без размера")


def _size_sort_value(size: str) -> tuple[int, int, str]:
    normalized = str(size or "").strip()
    if not normalized or normalized == "Без размера":
        return (999, 0, "")

    lower = normalized.lower()
    letter_sizes: dict[str, int] = {
        "xs": 1,
        "s": 2,
        "m": 3,
        "l": 4,
        "xl": 5,
        "xxl": 6,
        "xxxl": 7,
        "2xl": 6,
        "3xl": 7,
        "4xl": 8,
        "5xl": 9,
    }

    if lower in letter_sizes:
        return (1, letter_sizes[lower], lower)

    if "-" in normalized:
        try:
            first = int(normalized.split("-", 1)[0])
            return (2, first, lower)
        except Exception:
            pass

    digits = "".join(ch for ch in lower if ch.isdigit())
    if digits:
        try:
            return (2, int(digits), lower)
        except Exception:
            pass

    return (3, 0, lower)


def _compare_variant_key(variant: dict[str, Any]) -> tuple[Any, ...]:
    attrs = variant.get("attributes") or {}
    color = _variant_color(attrs)
    size = _variant_size(attrs)
    size_priority, size_number, size_label = _size_sort_value(size)
    try:
        pack_size = int(variant.get("pack_size") or 1)
    except Exception:
        pack_size = 1
    offer_id = str(variant.get("offer_id") or "")
    return (
        color.lower(),
        size_priority,
        size_number,
        size_label,
        pack_size,
        offer_id.lower(),
    )


def _month_title(month: str) -> str:
    try:
        year_str, month_str = str(month).split("-", 1)
        dt = datetime(int(year_str), int(month_str), 1)
        localized = dt.strftime("%B %Y")
        return localized[:1].upper() + localized[1:]
    except Exception:
        return str(month)


def _fmt_money(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except Exception:
        return 0.0


def _fmt_int(value: Any) -> int:
    try:
        return int(value or 0)
    except Exception:
        return 0


def build_warehouse_workbook(data: dict, *, order_window_days: int) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_name(f"Склад {data.get('store_name')}", "Склад")
    sheet.freeze_panes = "A5"

    row = 1
    title = f"Склад · {data.get('store_name') or 'Магазин'}"
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=15)
    cell.alignment = TEXT_ALIGN
    row += 1

    meta = [
        f"Наш склад: {data.get('warehouse_name') or '-'}",
        f"Период заказов: {order_window_days} дн.",
        f"Обновлено: {data.get('orders_updated_at') or '-'}",
    ]
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    meta_cell = sheet.cell(row=row, column=1, value=" · ".join(meta))
    meta_cell.alignment = TEXT_ALIGN
    row += 2

    headers = [
        "Товар",
        "Цвет",
        "Артикул",
        "Размер",
        "Уп.",
        "Неупак.",
        "Резерв",
        "Доступн",
        "OZON",
        "Готовим к продаже",
        "В заявках на поставку",
        "В поставках в пути",
        "Возвращаются от покупателей",
        f"Заказано за {('3 месяца' if order_window_days >= 90 else 'месяц' if order_window_days >= 30 else 'неделю')}",
        "Текущая цена",
        "Всего товара",
    ]
    _write_row(sheet, row, headers, fill=HEADER_FILL, font=HEADER_FONT)
    row += 1

    for product in data.get("products") or []:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        product_cell = sheet.cell(row=row, column=1, value=product.get("product_name") or "Товар")
        product_cell.font = TITLE_FONT
        product_cell.fill = SECTION_FILL
        product_cell.border = THIN_BORDER
        product_cell.alignment = TEXT_ALIGN
        row += 1

        sorted_variants = sorted(
            [variant for color_group in (product.get("colors") or []) for variant in (color_group.get("variants") or [])],
            key=_compare_variant_key,
        )

        product_data_start_row = row
        grouped_by_color: dict[str, list[dict]] = {}
        for variant in sorted_variants:
            display_color = _display_color(variant.get("color") or (variant.get("attributes") or {}).get("Цвет") or "")
            grouped_by_color.setdefault(display_color, []).append(variant)

        for color, variants in grouped_by_color.items():
            color_start_row = row
            current_size: str | None = None
            current_size_start_row: int | None = None

            for variant in variants:
                size_value = str(variant.get("size") or (variant.get("attributes") or {}).get("Размер") or "—")
                _write_row(
                    sheet,
                    row,
                    [
                        product.get("product_name") or "",
                        color or "—",
                        variant.get("offer_id") or "",
                        size_value,
                        variant.get("pack_size") or 1,
                        variant.get("warehouse_unpacked") or 0,
                        variant.get("warehouse_reserved") or 0,
                        variant.get("warehouse_available") or 0,
                        variant.get("ozon_available") or 0,
                        variant.get("ozon_ready_for_sale") or 0,
                        variant.get("ozon_requested_to_supply") or 0,
                        variant.get("ozon_in_transit") or 0,
                        variant.get("ozon_returning") or 0,
                        variant.get("ordered_units") or 0,
                        variant.get("current_price") or 0,
                        variant.get("total_units") or 0,
                    ],
                    alignment=CELL_ALIGN,
                )
                sheet.cell(row=row, column=1).alignment = TEXT_ALIGN
                sheet.cell(row=row, column=2).alignment = TEXT_ALIGN
                sheet.cell(row=row, column=3).alignment = TEXT_ALIGN

                if current_size is None:
                    current_size = size_value
                    current_size_start_row = row
                elif size_value != current_size:
                    if current_size_start_row is not None and row - current_size_start_row > 1:
                        sheet.merge_cells(
                            start_row=current_size_start_row,
                            start_column=4,
                            end_row=row - 1,
                            end_column=4,
                        )
                        sheet.cell(row=current_size_start_row, column=4).alignment = CELL_ALIGN
                    current_size = size_value
                    current_size_start_row = row

                row += 1

            if current_size_start_row is not None and row - current_size_start_row > 1:
                sheet.merge_cells(
                    start_row=current_size_start_row,
                    start_column=4,
                    end_row=row - 1,
                    end_column=4,
                )
                sheet.cell(row=current_size_start_row, column=4).alignment = CELL_ALIGN

            if row - color_start_row > 1:
                sheet.merge_cells(
                    start_row=color_start_row,
                    start_column=2,
                    end_row=row - 1,
                    end_column=2,
                )
                sheet.cell(row=color_start_row, column=2).alignment = CELL_ALIGN

        if row - product_data_start_row > 1:
            sheet.merge_cells(
                start_row=product_data_start_row,
                start_column=1,
                end_row=row - 1,
                end_column=1,
            )
            sheet.cell(row=product_data_start_row, column=1).alignment = TEXT_ALIGN

        row += 1

    _set_widths(
        sheet,
        {
            1: 28,
            2: 16,
            3: 28,
            4: 12,
            5: 8,
            6: 11,
            7: 10,
            8: 10,
            9: 10,
            10: 16,
            11: 18,
            12: 16,
            13: 20,
            14: 16,
            15: 14,
            16: 14,
        },
    )
    return workbook


def build_shipments_workbook(data: dict, *, order_window_days: int) -> Workbook:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    clusters = list(data.get("clusters") or [])
    if not clusters:
        sheet = workbook.create_sheet("Отправки")
        sheet["A1"] = "Нет данных для экспорта"
        return workbook

    order_label = "3 месяца" if order_window_days >= 90 else "месяц" if order_window_days >= 30 else "неделю"

    for cluster_index, cluster in enumerate(clusters, start=1):
        sheet = workbook.create_sheet(_safe_sheet_name(cluster.get("name") or f"Кластер {cluster_index}", f"Кластер {cluster_index}"))
        sheet.freeze_panes = "A5"
        row = 1

        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        title = sheet.cell(row=row, column=1, value=f"Кластер · {cluster.get('name') or f'Кластер {cluster_index}'}")
        title.font = Font(bold=True, size=15)
        title.alignment = TEXT_ALIGN
        row += 1

        summary = (
            f"OZON: {cluster.get('total_ozon_available') or 0} · "
            f"В пути: {cluster.get('total_in_pipeline') or 0} · "
            f"Заказано за {order_label}: {cluster.get('total_ordered_30d') or 0}"
        )
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        summary_cell = sheet.cell(row=row, column=1, value=summary)
        summary_cell.alignment = TEXT_ALIGN
        row += 2

        for warehouse in cluster.get("warehouses") or []:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            warehouse_cell = sheet.cell(row=row, column=1, value=f"{warehouse.get('name') or 'Склад'} · Ozon ID: {warehouse.get('ozon_id') or '-'}")
            warehouse_cell.font = TITLE_FONT
            warehouse_cell.fill = SECTION_FILL
            warehouse_cell.border = THIN_BORDER
            warehouse_cell.alignment = TEXT_ALIGN
            row += 1

            headers = [
                "Товар",
                "Артикул",
                "Цвет",
                "Размер",
                "Уп.",
                "Основной склад",
                "OZON",
                "В заявках на поставку",
                "В поставках в пути",
                "Возвращаются от покупателей",
                f"Заказано за {order_label}",
            ]
            _write_row(sheet, row, headers, fill=HEADER_FILL, font=HEADER_FONT)
            row += 1

            for product in warehouse.get("products") or []:
                grouped_by_color: dict[str, list[dict]] = {}
                sorted_variants = sorted(list(product.get("variants") or []), key=_compare_variant_key)
                for variant in sorted_variants:
                    color = _display_color((variant.get("attributes") or {}).get("Цвет") or "")
                    grouped_by_color.setdefault(color, []).append(variant)

                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
                product_cell = sheet.cell(row=row, column=1, value=product.get("product_name") or "Товар")
                product_cell.fill = SUBHEADER_FILL
                product_cell.font = HEADER_FONT
                product_cell.border = THIN_BORDER
                product_cell.alignment = TEXT_ALIGN
                row += 1

                product_data_start_row = row
                for color, variants in grouped_by_color.items():
                    group_start_row = row
                    current_size: str | None = None
                    current_size_start_row: int | None = None

                    for variant in variants:
                        attrs = variant.get("attributes") or {}
                        size_value = str(attrs.get("Размер") or "—")
                        _write_row(
                            sheet,
                            row,
                            [
                                product.get("product_name") or "",
                                variant.get("offer_id") or "",
                                color or "—",
                                size_value,
                                variant.get("pack_size") or 1,
                                variant.get("our_available") or 0,
                                variant.get("ozon_available") or 0,
                                variant.get("ozon_requested_to_supply") or 0,
                                variant.get("ozon_in_pipeline") or 0,
                                variant.get("ozon_returning") or 0,
                                variant.get("ordered_30d") or 0,
                            ],
                            alignment=CELL_ALIGN,
                        )
                        sheet.cell(row=row, column=1).alignment = TEXT_ALIGN
                        sheet.cell(row=row, column=2).alignment = TEXT_ALIGN
                        sheet.cell(row=row, column=3).alignment = TEXT_ALIGN

                        if current_size is None:
                            current_size = size_value
                            current_size_start_row = row
                        elif size_value != current_size:
                            if current_size_start_row is not None and row - current_size_start_row > 1:
                                sheet.merge_cells(
                                    start_row=current_size_start_row,
                                    start_column=4,
                                    end_row=row - 1,
                                    end_column=4,
                                )
                                sheet.cell(row=current_size_start_row, column=4).alignment = CELL_ALIGN
                            current_size = size_value
                            current_size_start_row = row
                        row += 1

                    if current_size_start_row is not None and row - current_size_start_row > 1:
                        sheet.merge_cells(
                            start_row=current_size_start_row,
                            start_column=4,
                            end_row=row - 1,
                            end_column=4,
                        )
                        sheet.cell(row=current_size_start_row, column=4).alignment = CELL_ALIGN

                    if row - group_start_row > 1:
                        sheet.merge_cells(
                            start_row=group_start_row,
                            start_column=3,
                            end_row=row - 1,
                            end_column=3,
                        )
                        sheet.cell(row=group_start_row, column=3).alignment = CELL_ALIGN

                if row - product_data_start_row > 1:
                    sheet.merge_cells(
                        start_row=product_data_start_row,
                        start_column=1,
                        end_row=row - 1,
                        end_column=1,
                    )
                    sheet.cell(row=product_data_start_row, column=1).alignment = TEXT_ALIGN

                row += 1

        _set_widths(
            sheet,
            {
                1: 28,
                2: 28,
                3: 14,
                4: 12,
                5: 8,
                6: 14,
                7: 10,
                8: 18,
                9: 16,
                10: 20,
                11: 16,
            },
        )

    return workbook


def build_closed_months_workbook(data: dict, *, year: int) -> Workbook:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = _safe_sheet_name(f"{year} сводка", "Сводка")
    summary_sheet.freeze_panes = "A5"

    store_name = str(data.get("store_name") or "Магазин")
    months = list(data.get("months") or [])

    row = 1
    summary_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    title_cell = summary_sheet.cell(row=row, column=1, value=f"Закрытые месяцы · {store_name} · {year}")
    title_cell.font = Font(bold=True, size=15)
    title_cell.alignment = TEXT_ALIGN
    row += 1

    summary_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    meta_cell = summary_sheet.cell(row=row, column=1, value=f"Месяцев в выгрузке: {len(months)}")
    meta_cell.alignment = TEXT_ALIGN
    row += 2

    summary_headers = [
        "Месяц",
        "Статус",
        "Покрытие",
        "Продано",
        "Выручка",
        "Валовая прибыль",
        "Чистая прибыль",
    ]
    _write_row(summary_sheet, row, summary_headers, fill=HEADER_FILL, font=HEADER_FONT)
    row += 1

    for item in months:
        month_data = item.get("month") or {}
        coverage_ratio = float(month_data.get("coverage_ratio") or 0)
        _write_row(
            summary_sheet,
            row,
            [
                _month_title(str(month_data.get("month") or "")),
                str(month_data.get("status") or ""),
                coverage_ratio,
                _fmt_int(month_data.get("sold_units")),
                _fmt_money(month_data.get("revenue_amount")),
                _fmt_money(month_data.get("gross_profit")),
                _fmt_money(month_data.get("net_profit")),
            ],
            alignment=CELL_ALIGN,
        )
        summary_sheet.cell(row=row, column=1).alignment = TEXT_ALIGN
        summary_sheet.cell(row=row, column=2).alignment = TEXT_ALIGN
        summary_sheet.cell(row=row, column=3).number_format = "0.0%"
        for col in (5, 6, 7):
            summary_sheet.cell(row=row, column=col).number_format = '#,##0.00 "₽"'
        row += 1

    _set_widths(summary_sheet, {1: 20, 2: 18, 3: 12, 4: 12, 5: 16, 6: 18, 7: 18})

    for index, item in enumerate(months, start=1):
        month_data = item.get("month") or {}
        offers = list(item.get("offers") or [])
        month_code = str(month_data.get("month") or f"{year}-{index:02d}")
        sheet = workbook.create_sheet(_safe_sheet_name(_month_title(month_code), f"Месяц {index}"))
        sheet.freeze_panes = "A9"

        row = 1
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = sheet.cell(row=row, column=1, value=f"{store_name} · {_month_title(month_code)}")
        cell.font = Font(bold=True, size=15)
        cell.alignment = TEXT_ALIGN
        row += 1

        coverage = float(month_data.get("coverage_ratio") or 0)
        meta_lines = [
            f"Статус: {month_data.get('status') or '-'}",
            f"Покрытие: {coverage:.1%}",
            f"Налоговый режим: {month_data.get('tax_mode_used') or '-'}",
            f"НДС: {month_data.get('vat_mode_used') or '-'}",
        ]
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        meta = sheet.cell(row=row, column=1, value=" · ".join(meta_lines))
        meta.alignment = TEXT_ALIGN
        row += 2

        metric_rows = [
            ("Продано", _fmt_int(month_data.get("sold_units")), "Штук, которые Ozon закрыл как продажи"),
            ("Возвраты", _fmt_int(month_data.get("returned_units")), "Штук, закрытых как возвраты"),
            ("Выручка", _fmt_money(month_data.get("revenue_amount")), "Реализация минус возвраты"),
            ("Себестоимость", _fmt_money(month_data.get("cogs")), "Историческая себестоимость месяца"),
            ("Валовая прибыль", _fmt_money(month_data.get("gross_profit")), "Выручка без НДС минус себестоимость"),
            ("Комиссия Ozon", _fmt_money(month_data.get("ozon_commission")), "По отчету реализации"),
            ("Логистика Ozon", _fmt_money(month_data.get("ozon_logistics")), "Доставка и возвратная логистика"),
            ("Услуги Ozon", _fmt_money(month_data.get("ozon_services")), "Маркетинг, хранение и сервисы"),
            ("Эквайринг Ozon", _fmt_money(month_data.get("ozon_acquiring")), "Платежный эквайринг"),
            ("Прочие расходы", _fmt_money(month_data.get("ozon_other_expenses")), "Прочие расходы Ozon"),
            ("Бонусы и софинансирование", _fmt_money(month_data.get("ozon_incentives")), "Партнерские выплаты и бонусы"),
            ("Корректировки Ozon", _fmt_money(month_data.get("ozon_adjustments_net")), "Компенсации минус декомпенсации"),
            ("Прибыль до налога", _fmt_money(month_data.get("profit_before_tax")), "После расходов Ozon, до налога"),
            ("Налог", _fmt_money(month_data.get("tax_amount")), "По исторической налоговой схеме"),
            ("Чистая прибыль", _fmt_money(month_data.get("net_profit")), "Итог после всех расходов и налога"),
        ]
        _write_row(sheet, row, ["Показатель", "Значение", "Описание"], fill=HEADER_FILL, font=HEADER_FONT)
        row += 1
        for label, value, description in metric_rows:
            _write_row(sheet, row, [label, value, description], alignment=CELL_ALIGN)
            sheet.cell(row=row, column=1).alignment = TEXT_ALIGN
            sheet.cell(row=row, column=3).alignment = TEXT_ALIGN
            if isinstance(value, float):
                sheet.cell(row=row, column=2).number_format = '#,##0.00 "₽"'
            row += 1

        row += 1
        offer_headers = [
            "Артикул",
            "Название",
            "Продано",
            "Возвраты",
            "Выручка",
            "Себестоимость",
            "Валовая прибыль",
            "Прибыль до налога",
            "Налог",
            "Чистая прибыль",
            "Себестоимость заполнена",
        ]
        _write_row(sheet, row, offer_headers, fill=SUBHEADER_FILL, font=HEADER_FONT)
        row += 1

        for offer in offers:
            _write_row(
                sheet,
                row,
                [
                    str(offer.get("offer_id") or ""),
                    str(offer.get("title") or ""),
                    _fmt_int(offer.get("sold_units")),
                    _fmt_int(offer.get("returned_units")),
                    _fmt_money(offer.get("revenue_amount")),
                    _fmt_money(offer.get("cogs")) if offer.get("cogs") is not None else "—",
                    _fmt_money(offer.get("gross_profit")) if offer.get("gross_profit") is not None else "—",
                    _fmt_money(offer.get("profit_before_tax")) if offer.get("profit_before_tax") is not None else "—",
                    _fmt_money(offer.get("tax_amount")) if offer.get("tax_amount") is not None else "—",
                    _fmt_money(offer.get("net_profit")) if offer.get("net_profit") is not None else "—",
                    "Да" if bool(offer.get("has_cost")) else "Нет",
                ],
                alignment=CELL_ALIGN,
            )
            sheet.cell(row=row, column=1).alignment = TEXT_ALIGN
            sheet.cell(row=row, column=2).alignment = TEXT_ALIGN
            for col in (5, 6, 7, 8, 9, 10):
                if isinstance(sheet.cell(row=row, column=col).value, float):
                    sheet.cell(row=row, column=col).number_format = '#,##0.00 "₽"'
            row += 1

        _set_widths(
            sheet,
            {
                1: 26,
                2: 34,
                3: 11,
                4: 11,
                5: 15,
                6: 15,
                7: 16,
                8: 16,
                9: 14,
                10: 16,
                11: 16,
            },
        )

    return workbook


def export_file_path(kind: str, store_id: int) -> tuple[str, str]:
    root = ensure_export_root_dir()
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    file_name = f"{kind}_store_{store_id}_{timestamp}.xlsx"
    return os.path.join(root, file_name), file_name
