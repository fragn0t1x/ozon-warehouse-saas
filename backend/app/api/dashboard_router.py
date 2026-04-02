import asyncio
import csv
import io
import zipfile
from datetime import date, datetime, timedelta, timezone
from statistics import median
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from loguru import logger

from app.core.dependencies import get_current_user
from app.database import SessionLocal
from app.models.product import Product
from app.models.ozon_warehouse import OzonStock
from app.models.store import Store
from app.models.supply import Supply
from app.models.user import User
from app.models.variant import Variant
from app.models.warehouse import Warehouse, WarehouseStock
from app.services.admin_notifications import get_recent_admin_events
from app.services.cabinet_access import get_cabinet_owner_id
from app.services.dashboard_finance_service import DashboardFinanceService
from app.services.dashboard_sales_service import DashboardSalesService
from app.services.dashboard_unit_economics_service import DashboardUnitEconomicsService
from app.services.economics_history_service import EconomicsHistoryService
from app.services.ozon.client import OzonClient
from app.services.ozon.report_schema_guard import missing_required_column_groups, notify_ozon_report_columns_changed
from app.services.ozon.report_service import OzonReportService
from app.services.ozon.report_snapshot_service import OzonReportSnapshotService
from app.services.settings_service import SettingsService
from app.utils.encryption import decrypt_api_key
from app.services.supply_reservation_wait import get_supply_reservation_wait_map
from app.services.sync_dispatcher import (
    enqueue_finance_snapshot_sync,
    enqueue_report_snapshot_sync,
    enqueue_stocks_sync,
)

router = APIRouter(prefix="/dashboard", tags=["dashboard"])

RETURNS_REQUIRED_COLUMNS: tuple[tuple[str, ...], ...] = (
    ("Количество возвращаемых товаров", "Returned quantity", "returned quantity", "Quantity of returned goods"),
)


async def get_db():
    async with SessionLocal() as session:
        yield session


def _iso_or_none(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def _shift_month(source: date, months: int) -> date:
    month_index = (source.month - 1) + months
    year = source.year + month_index // 12
    month = month_index % 12 + 1
    return date(year, month, 1)


def _same_day_previous_month(source: date) -> date:
    previous_month_start = _shift_month(source.replace(day=1), -1)
    current_month_start = source.replace(day=1)
    last_day_previous_month = (current_month_start - timedelta(days=1)).day
    return previous_month_start.replace(day=min(source.day, last_day_previous_month))


def _to_float(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _normalize_header(value: str) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _pick_value(row: dict[str, str], *aliases: str) -> str:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(_normalize_header(alias))
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _decode_bytes(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode("utf-8", errors="replace")


def _iter_xlsx_rows(raw_bytes: bytes) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    header = _extract_returns_xlsx_headers(raw_bytes)
    if not header:
        return rows
    with io.BytesIO(raw_bytes) as buffer:
        workbook = load_workbook(buffer, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            header_seen = False
            for raw_row in worksheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value).strip() for value in raw_row]
                if not any(values):
                    continue
                normalized_values = {_normalize_header(v) for v in values if v}
                if not header_seen:
                    if [_normalize_header(v) for v in values if v] == [_normalize_header(v) for v in header if v]:
                        header_seen = True
                    elif set(_normalize_header(v) for v in header if v).intersection(normalized_values):
                        header_seen = True
                    continue
                padded = values + [""] * max(0, len(header) - len(values))
                row = {
                    str(header[index]).strip(): str(padded[index]).strip()
                    for index in range(len(header))
                    if str(header[index]).strip()
                }
                if any(row.values()):
                    rows.append(row)
    return rows


def _extract_returns_xlsx_headers(raw_bytes: bytes) -> list[str]:
    header_aliases = {
        "количество возвращаемых товаров",
        "статус возврата",
        "ozon sku id",
        "артикул товара",
        "номер отправления",
    }
    with io.BytesIO(raw_bytes) as buffer:
        workbook = load_workbook(buffer, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            for raw_row in worksheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value).strip() for value in raw_row]
                if not any(values):
                    continue
                normalized_values = {_normalize_header(v) for v in values if v}
                if header_aliases.intersection(normalized_values):
                    return [value for value in values if value]
    return []


def _parse_returns_units(raw_bytes: bytes) -> int:
    total_units = 0.0
    parsed_any_rows = False

    quantity_aliases = (
        "Количество возвращаемых товаров",
        "Returned quantity",
        "returned quantity",
        "Quantity of returned goods",
    )

    if zipfile.is_zipfile(io.BytesIO(raw_bytes)):
        try:
            raw_rows = _iter_xlsx_rows(raw_bytes)
            for raw_row in raw_rows:
                row = {str(key).strip(): str(value).strip() for key, value in raw_row.items() if key}
                if not any(row.values()):
                    continue
                quantity = _pick_value(row, *quantity_aliases)
                total_units += _to_float(quantity)
            return int(round(total_units))
        except Exception as exc:
            logger.warning("Failed to parse returns XLSX report, falling back to text parser: {}", exc)

    text = _decode_bytes(raw_bytes)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;	")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    for raw_row in reader:
        row = {str(key).strip(): str(value).strip() for key, value in raw_row.items() if key}
        if not any(row.values()):
            continue
        parsed_any_rows = True
        quantity = _pick_value(row, *quantity_aliases)
        total_units += _to_float(quantity)

    if not parsed_any_rows:
        raise RuntimeError("Returns report did not contain readable rows")

    return int(round(total_units))


async def _fetch_store_returns_units_for_period(store: dict[str, object], *, date_from: date, date_to: date) -> dict[str, object]:
    client_id = str(store.get("client_id") or "").strip()
    api_key_encrypted = store.get("api_key_encrypted")
    store_name = str(store.get("name") or "Магазин")
    if not client_id or not api_key_encrypted:
        return {"available": False, "units": 0}

    client = OzonClient(
        client_id,
        decrypt_api_key(api_key_encrypted),
        store_name=store_name,
        emit_notifications=False,
    )
    try:
        report_code = await client.create_returns_report(
            date_from=f"{date_from.isoformat()}T00:00:00.000Z",
            date_to=f"{date_to.isoformat()}T23:59:59.999Z",
            status="MovingToOzon",
            language="DEFAULT",
        )
        if not report_code:
            raise RuntimeError("Ozon did not return report code for returns report")

        report_info: dict[str, object] = {}
        for _ in range(12):
            report_info = await client.get_report_info(report_code)
            status = str(report_info.get("status") or "").lower()
            if status == "success" and report_info.get("file"):
                break
            if status in {"failed", "error", "cancelled"}:
                raise RuntimeError(f"Returns report failed with status={report_info.get('status')}: {report_info.get('error') or ''}")
            await asyncio.sleep(2)
        else:
            raise RuntimeError(f"Returns report was not ready in time for code={report_code}")

        file_url = str(report_info.get("file") or "").strip()
        if not file_url:
            raise RuntimeError(f"Returns report {report_code} does not contain file URL")

        raw_bytes = await client.download_report_file(file_url)
        headers: list[str] = []
        try:
            if zipfile.is_zipfile(io.BytesIO(raw_bytes)):
                headers = _extract_returns_xlsx_headers(raw_bytes)
            else:
                text = _decode_bytes(raw_bytes)
                sample = text[:4096]
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                except csv.Error:
                    dialect = csv.excel
                    dialect.delimiter = ";"
                headers = [str(header).strip() for header in (csv.DictReader(io.StringIO(text), dialect=dialect).fieldnames or []) if header]
        except Exception:
            headers = []
        missing_groups = missing_required_column_groups(headers, RETURNS_REQUIRED_COLUMNS)
        if missing_groups:
            await notify_ozon_report_columns_changed(
                endpoint="/v2/report/returns/create",
                client_id=client_id,
                report_name=f"dashboard_returns:{store_name}",
                required_groups=missing_groups,
                actual_headers=headers,
                payload={
                    "date_from": date_from.isoformat(),
                    "date_to": date_to.isoformat(),
                },
            )
            raise RuntimeError("Returns report is missing required columns")
        units = _parse_returns_units(raw_bytes)
        return {"available": True, "units": units, "report_code": report_code}
    except Exception as exc:
        logger.warning(
            "Failed to fetch returns units for store={} client_id={} period={}..{}: {}",
            store_name,
            client_id,
            date_from.isoformat(),
            date_to.isoformat(),
            exc,
        )
        return {"available": False, "units": 0, "error": str(exc)}
    finally:
        await client.close()


async def _avg_delivery_days(db: AsyncSession, store_id: int, storage_warehouse_id: int | None) -> int:
    stmt = select(Supply).where(
        Supply.store_id == store_id,
        Supply.timeslot_from.is_not(None),
        func.coalesce(Supply.acceptance_at_storage_at, Supply.completed_at).is_not(None),
        func.coalesce(Supply.acceptance_at_storage_at, Supply.completed_at) >= Supply.timeslot_from,
    ).order_by(func.coalesce(Supply.acceptance_at_storage_at, Supply.completed_at).desc()).limit(20)

    if storage_warehouse_id:
        stmt = stmt.where(Supply.storage_warehouse_id == storage_warehouse_id)

    result = await db.execute(stmt)
    supplies = result.scalars().all()
    if not supplies:
        return 2

    deltas: list[int] = []
    for supply in supplies:
        try:
            arrival_at = supply.acceptance_at_storage_at or supply.completed_at
            if arrival_at is None:
                continue
            delta = (arrival_at.date() - supply.timeslot_from.date()).days
            deltas.append(max(0, delta))
        except Exception:
            continue

    if not deltas:
        return 2

    return max(1, round(median(deltas)))


def _resolve_eta_iso(supply: Supply, avg_days: int | None = None) -> str | None:
    if supply.eta_date:
        return supply.eta_date.isoformat()
    if supply.timeslot_from and avg_days:
        return (supply.timeslot_from.date() + timedelta(days=avg_days)).isoformat()
    return None


@router.get("/summary")
async def get_dashboard_summary(
    store_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    now = datetime.now()
    tomorrow = now + timedelta(days=1)
    cabinet_owner_id = get_cabinet_owner_id(current_user)
    settings = await SettingsService(db).get_settings(current_user.id)

    store_rows_result = await db.execute(
        select(
            Store.id,
            Store.name,
            Store.client_id,
            Store.api_key_encrypted,
            Store.economics_vat_mode,
            Store.economics_tax_mode,
            Store.economics_tax_rate,
            Store.economics_default_sale_price_gross,
        ).where(Store.user_id == cabinet_owner_id)
    )
    all_user_stores = [
        {
            "id": store_id,
            "name": store_name,
            "client_id": client_id,
            "api_key_encrypted": api_key_encrypted,
            "economics_vat_mode": economics_vat_mode,
            "economics_tax_mode": economics_tax_mode,
            "economics_tax_rate": economics_tax_rate,
            "economics_default_sale_price_gross": economics_default_sale_price_gross,
        }
        for (
            store_id,
            store_name,
            client_id,
            api_key_encrypted,
            economics_vat_mode,
            economics_tax_mode,
            economics_tax_rate,
            economics_default_sale_price_gross,
        ) in store_rows_result.all()
    ]
    all_store_ids = {int(store["id"]) for store in all_user_stores}
    if store_id is not None and store_id not in all_store_ids:
        raise HTTPException(status_code=404, detail="Магазин не найден")

    user_stores = [
        store for store in all_user_stores
        if store_id is None or int(store["id"]) == store_id
    ]
    store_ids = [int(store["id"]) for store in user_stores]
    store_name_by_id = {int(store["id"]): str(store["name"]) for store in user_stores}

    store_count = len(store_ids)

    product_count = 0
    variant_count = 0
    if store_ids:
        product_count = await db.scalar(
            select(func.count(Product.id))
            .where(Product.store_id.in_(store_ids))
        ) or 0

        variant_count = await db.scalar(
            select(func.count(Variant.id))
            .join(Product, Product.id == Variant.product_id)
            .where(Product.store_id.in_(store_ids))
        ) or 0

    uses_shared_warehouse = bool(settings and settings.warehouse_mode == "shared")

    warehouse_count_stmt = select(func.count(Warehouse.id)).where(Warehouse.user_id == cabinet_owner_id)
    if store_id is not None:
        if uses_shared_warehouse and settings and settings.shared_warehouse_id:
            warehouse_count_stmt = warehouse_count_stmt.where(Warehouse.id == settings.shared_warehouse_id)
        else:
            warehouse_count_stmt = warehouse_count_stmt.where(Warehouse.store_id == store_id)

    warehouse_count = await db.scalar(warehouse_count_stmt) or 0

    status_rows = []
    recent_supplies = []
    recent_supplies_by_store: list[dict] = []
    waiting_for_stock_supplies = 0

    if store_ids:
        waiting_stmt = select(Supply.id).where(
            Supply.store_id.in_(store_ids),
            Supply.status == "READY_TO_SUPPLY",
            Supply.reserved_at.is_(None),
        )
        waiting_supply_ids = [row[0] for row in (await db.execute(waiting_stmt)).all()]
        waiting_map = await get_supply_reservation_wait_map(waiting_supply_ids)
        waiting_for_stock_supplies = len(waiting_map)

        status_result = await db.execute(
            select(Supply.status, func.count(Supply.id))
            .where(Supply.store_id.in_(store_ids))
            .group_by(Supply.status)
        )
        status_rows = status_result.all()

        recent_result = await db.execute(
            select(Supply, Store.name)
            .join(Store, Store.id == Supply.store_id)
            .where(Store.user_id == cabinet_owner_id)
            .order_by(
                Supply.timeslot_from.desc().nullslast(),
                Supply.created_at.desc(),
            )
            .limit(min(max(store_count * 18, 36), 180))
        )
        recent_rows = recent_result.all()
        recent_wait_map = await get_supply_reservation_wait_map([supply.id for supply, _store_name in recent_rows])
        avg_cache: dict[tuple[int, int | None], int] = {}
        grouped_recent: dict[int, list[dict]] = {}
        for supply, store_name in recent_rows:
            store_bucket = grouped_recent.setdefault(int(supply.store_id), [])
            if len(store_bucket) >= 6:
                continue
            avg_key = (int(supply.store_id), supply.storage_warehouse_id)
            if avg_key not in avg_cache:
                avg_cache[avg_key] = await _avg_delivery_days(db, int(supply.store_id), supply.storage_warehouse_id)
            store_bucket.append(
                {
                    "id": supply.id,
                    "order_number": supply.order_number,
                    "status": supply.status,
                    "reservation_waiting_for_stock": supply.id in recent_wait_map,
                    "reservation_wait_message": recent_wait_map.get(supply.id, {}).get("message"),
                    "store_id": int(supply.store_id),
                    "store_name": store_name,
                    "eta_date": _resolve_eta_iso(supply, avg_cache.get(avg_key)),
                    "timeslot_from": _iso_or_none(supply.timeslot_from),
                }
            )

        recent_supplies_by_store = [
            {
                "store_id": int(store["id"]),
                "store_name": store["name"],
                "items": grouped_recent.get(int(store["id"]), []),
            }
            for store in user_stores
        ]
        recent_supplies = [
            item
            for group in recent_supplies_by_store
            for item in group["items"]
        ][:12]

    status_counts = {status: count for status, count in status_rows}
    active_statuses = {
        "DATA_FILLING",
        "READY_TO_SUPPLY",
        "ACCEPTED_AT_SUPPLY_WAREHOUSE",
        "IN_TRANSIT",
        "ACCEPTANCE_AT_STORAGE_WAREHOUSE",
        "REPORTS_CONFIRMATION_AWAITING",
    }
    active_supplies = sum(
        count for status, count in status_counts.items() if status in active_statuses
    )

    today_supplies = 0
    if store_ids:
        today_supplies = await db.scalar(
            select(func.count(Supply.id))
            .where(
                Supply.store_id.in_(store_ids),
                Supply.timeslot_from >= now,
                Supply.timeslot_from < tomorrow,
            )
        ) or 0

    warehouses_stmt = select(Warehouse.id, Warehouse.name, Warehouse.store_id).where(Warehouse.user_id == cabinet_owner_id)
    if store_id is not None:
        if uses_shared_warehouse and settings and settings.shared_warehouse_id:
            warehouses_stmt = warehouses_stmt.where(Warehouse.id == settings.shared_warehouse_id)
        else:
            warehouses_stmt = warehouses_stmt.where(Warehouse.store_id == store_id)

    warehouses_result = await db.execute(warehouses_stmt)
    warehouse_rows = warehouses_result.all()
    warehouse_meta_by_id = {
        warehouse_id: {
            "warehouse_id": int(warehouse_id),
            "warehouse_name": warehouse_name,
            "store_id": int(store_id) if store_id is not None else None,
            "store_name": store_name_by_id.get(int(store_id)) if store_id is not None else None,
        }
        for warehouse_id, warehouse_name, store_id in warehouse_rows
    }

    warehouse_agg_stmt = (
        select(
            WarehouseStock.warehouse_id,
            func.coalesce(func.sum(WarehouseStock.unpacked_quantity), 0),
            func.coalesce(func.sum(WarehouseStock.packed_quantity), 0),
            func.coalesce(func.sum(WarehouseStock.packed_quantity * func.coalesce(Variant.pack_size, 1)), 0),
            func.coalesce(func.sum(WarehouseStock.reserved_quantity), 0),
            func.max(WarehouseStock.updated_at),
        )
        .join(Warehouse, Warehouse.id == WarehouseStock.warehouse_id)
        .join(Variant, Variant.id == WarehouseStock.variant_id)
        .where(Warehouse.user_id == cabinet_owner_id)
        .group_by(WarehouseStock.warehouse_id)
    )
    if store_id is not None:
        if uses_shared_warehouse and settings and settings.shared_warehouse_id:
            warehouse_agg_stmt = warehouse_agg_stmt.where(Warehouse.id == settings.shared_warehouse_id)
        else:
            warehouse_agg_stmt = warehouse_agg_stmt.where(Warehouse.store_id == store_id)

    warehouse_agg_result = await db.execute(warehouse_agg_stmt)
    warehouse_agg_rows = warehouse_agg_result.all()
    warehouse_stats_map = {}
    for warehouse_id, unpacked, packed_boxes, packed_units, reserved, updated_at in warehouse_agg_rows:
        warehouse_stats_map[int(warehouse_id)] = {
            "unpacked_units": int(unpacked or 0),
            "packed_boxes": int(packed_boxes or 0),
            "packed_units": int(packed_units or 0),
            "reserved_units": int(reserved or 0),
            "available_units": int((unpacked or 0) + (packed_units or 0) - (reserved or 0)),
            "updated_at": _iso_or_none(updated_at),
        }

    ozon_overall_result = await db.execute(
        select(
            func.coalesce(func.sum(OzonStock.available_to_sell), 0),
            func.coalesce(func.sum(OzonStock.in_supply), 0),
            func.coalesce(func.sum(OzonStock.in_transit), 0),
            func.max(OzonStock.updated_at),
        )
        .join(Variant, Variant.id == OzonStock.variant_id)
        .join(Product, Product.id == Variant.product_id)
        .join(Store, Store.id == Product.store_id)
        .where(Store.user_id == cabinet_owner_id)
    )
    ozon_available, ozon_in_supply, ozon_in_transit, ozon_updated_at = ozon_overall_result.one()

    ozon_by_store_result = await db.execute(
        select(
            Store.id,
            func.coalesce(func.sum(OzonStock.available_to_sell), 0),
            func.coalesce(func.sum(OzonStock.in_supply), 0),
            func.coalesce(func.sum(OzonStock.in_transit), 0),
            func.max(OzonStock.updated_at),
        )
        .join(Product, Product.store_id == Store.id)
        .join(Variant, Variant.product_id == Product.id)
        .join(OzonStock, OzonStock.variant_id == Variant.id)
        .where(Store.user_id == cabinet_owner_id)
        .group_by(Store.id)
    )
    ozon_store_stats_map = {
        int(store_id): {
            "available_to_sell": int(available_to_sell or 0),
            "in_supply": int(in_supply or 0),
            "in_transit": int(in_transit or 0),
            "updated_at": _iso_or_none(updated_at),
        }
        for store_id, available_to_sell, in_supply, in_transit, updated_at in ozon_by_store_result.all()
    }

    total_unpacked = sum(item["unpacked_units"] for item in warehouse_stats_map.values())
    total_packed_boxes = sum(item["packed_boxes"] for item in warehouse_stats_map.values())
    total_packed_units = sum(item["packed_units"] for item in warehouse_stats_map.values())
    total_reserved = sum(item["reserved_units"] for item in warehouse_stats_map.values())
    total_available = sum(item["available_units"] for item in warehouse_stats_map.values())
    latest_warehouse_update = max(
        (item["updated_at"] for item in warehouse_stats_map.values() if item.get("updated_at")),
        default=None,
    )

    stock_source = "warehouse"
    stock_note = "По внутреннему складу."
    stock_updated_at = latest_warehouse_update

    if not warehouse_stats_map:
        total_unpacked = int(ozon_available or 0)
        total_packed_boxes = 0
        total_packed_units = int((ozon_in_supply or 0) + (ozon_in_transit or 0))
        total_reserved = 0
        total_available = int(ozon_available or 0)
        stock_source = "ozon"
        stock_note = "Пока нет ручного прихода на внутренний склад, поэтому показываем актуальные остатки OZON."
        stock_updated_at = _iso_or_none(ozon_updated_at)
    elif store_id is not None:
        if uses_shared_warehouse:
            stock_note = "Показываем остатки общего склада кабинета, который сейчас использует активный магазин."
        else:
            stock_note = "Показываем остатки внутреннего склада активного магазина."

    shared_warehouse_meta = warehouse_meta_by_id.get(settings.shared_warehouse_id) if settings and settings.shared_warehouse_id else None
    shared_warehouse_stats = warehouse_stats_map.get(settings.shared_warehouse_id) if settings and settings.shared_warehouse_id else None

    warehouse_breakdown = []
    for warehouse_id, meta in sorted(
        warehouse_meta_by_id.items(),
        key=lambda item: ((item[1].get("store_name") or ""), item[1]["warehouse_name"]),
    ):
        stats = warehouse_stats_map.get(warehouse_id, {})
        warehouse_breakdown.append(
            {
                **meta,
                "unpacked_units": int(stats.get("unpacked_units", 0)),
                "packed_boxes": int(stats.get("packed_boxes", 0)),
                "packed_units": int(stats.get("packed_units", 0)),
                "reserved_units": int(stats.get("reserved_units", 0)),
                "available_units": int(stats.get("available_units", 0)),
                "updated_at": stats.get("updated_at"),
            }
        )

    stock_by_store = []
    orders_snapshot_by_store: dict[int, dict | None] = {}
    orders_totals_by_store: dict[int, int] = {}
    orders_period_from_values: list[str] = []
    orders_period_to_values: list[str] = []
    orders_updated_values: list[str] = []
    orders_stores_covered = 0
    for store in user_stores:
        snapshot = await OzonReportSnapshotService.get_cached_snapshot_for_client(
            client_id=store["client_id"],
            kind="postings",
        )
        orders_snapshot_by_store[int(store["id"])] = snapshot
        orders_totals_by_store[int(store["id"])] = OzonReportSnapshotService.get_total_postings_order_units(snapshot)
        if snapshot:
            orders_stores_covered += 1
            filters = snapshot.get("filters") or {}
            if filters.get("processed_at_from"):
                orders_period_from_values.append(filters["processed_at_from"])
            if filters.get("processed_at_to"):
                orders_period_to_values.append(filters["processed_at_to"])
            if snapshot.get("refreshed_at"):
                orders_updated_values.append(snapshot["refreshed_at"])

    for store in user_stores:
        store_id = int(store["id"])
        per_store_warehouse_meta = next(
            (meta for meta in warehouse_meta_by_id.values() if meta.get("store_id") == store_id),
            None,
        )
        per_store_warehouse_stats = warehouse_stats_map.get(per_store_warehouse_meta["warehouse_id"]) if per_store_warehouse_meta else None
        effective_warehouse_meta = shared_warehouse_meta if uses_shared_warehouse else per_store_warehouse_meta
        effective_warehouse_stats = shared_warehouse_stats if uses_shared_warehouse else per_store_warehouse_stats
        ozon_stats = ozon_store_stats_map.get(store_id, {})

        stock_by_store.append(
            {
                "store_id": store_id,
                "store_name": store["name"],
                "warehouse_scope": "shared" if uses_shared_warehouse else "per_store",
                "warehouse_id": effective_warehouse_meta["warehouse_id"] if effective_warehouse_meta else None,
                "warehouse_name": effective_warehouse_meta["warehouse_name"] if effective_warehouse_meta else None,
                "warehouse_unpacked_units": int((effective_warehouse_stats or {}).get("unpacked_units", 0)),
                "warehouse_packed_boxes": int((effective_warehouse_stats or {}).get("packed_boxes", 0)),
                "warehouse_packed_units": int((effective_warehouse_stats or {}).get("packed_units", 0)),
                "warehouse_reserved_units": int((effective_warehouse_stats or {}).get("reserved_units", 0)),
                "warehouse_available_units": int((effective_warehouse_stats or {}).get("available_units", 0)),
                "warehouse_updated_at": (effective_warehouse_stats or {}).get("updated_at"),
                "ozon_available_units": int(ozon_stats.get("available_to_sell", 0)),
                "ozon_in_transit_units": int(ozon_stats.get("in_transit", 0)),
                "ozon_updated_at": ozon_stats.get("updated_at"),
                "ordered_30d_units": int(orders_totals_by_store.get(store_id, 0)),
                "orders_updated_at": (orders_snapshot_by_store.get(store_id) or {}).get("refreshed_at"),
            }
        )

    sales_summary = await DashboardSalesService().build_fbo_sales_summary(
        stores=user_stores,
        period_days=30,
    )
    finance_summary = await DashboardFinanceService().build_finance_summary(
        stores=user_stores,
        period_days=62,
    )

    today = now.date()
    current_month_start = today.replace(day=1)
    previous_month_start = _shift_month(current_month_start, -1)
    previous_month_same_day = _same_day_previous_month(today)

    current_returns_tasks = [
        asyncio.create_task(_fetch_store_returns_units_for_period(store, date_from=current_month_start, date_to=today))
        for store in user_stores
    ]
    previous_returns_tasks = [
        asyncio.create_task(_fetch_store_returns_units_for_period(store, date_from=previous_month_start, date_to=previous_month_same_day))
        for store in user_stores
    ]

    current_returns_results = await asyncio.gather(*current_returns_tasks, return_exceptions=True) if current_returns_tasks else []
    previous_returns_results = await asyncio.gather(*previous_returns_tasks, return_exceptions=True) if previous_returns_tasks else []

    returned_units = 0
    previous_returned_units = 0
    returned_units_covered = 0
    previous_returned_units_covered = 0

    for result in current_returns_results:
        if isinstance(result, Exception):
            continue
        if result.get("available"):
            returned_units += int(result.get("units") or 0)
            returned_units_covered += 1

    for result in previous_returns_results:
        if isinstance(result, Exception):
            continue
        if result.get("available"):
            previous_returned_units += int(result.get("units") or 0)
            previous_returned_units_covered += 1

    finance_summary_current_month = finance_summary.get("current_month_to_date") or {}
    finance_summary_current_month.update({
        "returned_units": returned_units if returned_units_covered > 0 else None,
        "previous_returned_units": previous_returned_units if previous_returned_units_covered > 0 else None,
        "delta_returned_units": (returned_units - previous_returned_units) if returned_units_covered > 0 and previous_returned_units_covered > 0 else None,
        "returned_units_available": returned_units_covered > 0,
        "returned_units_delta_available": returned_units_covered > 0 and previous_returned_units_covered > 0,
    })
    finance_summary["current_month_to_date"] = finance_summary_current_month

    unit_economics_summary = await DashboardUnitEconomicsService().build_summary(
        db=db,
        user_id=cabinet_owner_id,
        stores=user_stores,
        warehouse_mode=settings.warehouse_mode if settings else "shared",
    )

    return {
        "stores": store_count,
        "products": product_count,
        "variants": variant_count,
        "warehouses": warehouse_count,
        "today_supplies": today_supplies,
        "active_supplies": active_supplies,
        "waiting_for_stock_supplies": waiting_for_stock_supplies,
        "status_counts": status_counts,
        "warehouse_mode": settings.warehouse_mode if settings else "shared",
        "packing_mode": settings.packing_mode if settings else "simple",
        "discrepancy_mode": settings.discrepancy_mode if settings else "loss",
        "is_first_login": False if current_user.is_admin else settings.is_first_login,
        "shipments_start_date": settings.shipments_start_date.isoformat() if settings and settings.shipments_start_date else None,
        "shipments_accounting_enabled": settings.shipments_accounting_enabled if settings else False,
        "shipments_accounting_enabled_at": (
            settings.shipments_accounting_enabled_at.isoformat()
            if settings and settings.shipments_accounting_enabled_at
            else None
        ),
        "stock": {
            "unpacked_units": total_unpacked,
            "packed_boxes": total_packed_boxes,
            "packed_units": total_packed_units,
            "reserved_units": total_reserved,
            "available_units": total_available,
            "source": stock_source,
            "note": stock_note,
            "updated_at": stock_updated_at,
            "ordered_30d_units": int(sum(orders_totals_by_store.values())),
            "orders_updated_at": max(orders_updated_values, default=None),
            "orders_period_from": min(orders_period_from_values, default=None),
            "orders_period_to": max(orders_period_to_values, default=None),
            "orders_stores_covered": orders_stores_covered,
            "orders_stores_missing": max(len(user_stores) - orders_stores_covered, 0),
        },
        "warehouse_breakdown": warehouse_breakdown,
        "stock_by_store": stock_by_store,
        "sales": sales_summary,
        "finance": finance_summary,
        "unit_economics": unit_economics_summary,
        "recent_supplies": recent_supplies,
        "recent_supplies_by_store": recent_supplies_by_store,
        "admin_events": await get_recent_admin_events(8) if current_user.is_admin else [],
    }


@router.post("/refresh-commercial")
async def refresh_dashboard_commercial_data(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    store_rows_result = await db.execute(
        select(Store.id, Store.name)
        .where(
            Store.user_id == get_cabinet_owner_id(current_user),
            Store.is_active == True,  # noqa: E712
        )
        .order_by(Store.id)
    )
    stores = [
        {"id": store_id, "name": store_name}
        for store_id, store_name in store_rows_result.all()
    ]

    if not stores:
        return {
            "queued": 0,
            "store_names": [],
            "queued_tasks": {
                "report_snapshots": 0,
                "finance_snapshots": 0,
                "stocks": 0,
            },
            "message": "У пользователя нет активных магазинов",
        }

    queued_report_tasks = 0
    queued_finance_tasks = 0
    queued_stocks_tasks = 0

    for store in stores:
        enqueue_report_snapshot_sync(store["id"])
        queued_report_tasks += 1

        enqueue_finance_snapshot_sync(store["id"])
        queued_finance_tasks += 1

        enqueue_stocks_sync(store["id"])
        queued_stocks_tasks += 1

    return {
        "queued": len(stores),
        "store_names": [store["name"] for store in stores],
        "queued_tasks": {
            "report_snapshots": queued_report_tasks,
            "finance_snapshots": queued_finance_tasks,
            "stocks": queued_stocks_tasks,
        },
        "message": "Обновление продаж, финансов и остатков поставлено в очередь",
    }


@router.get("/unit-economics")
async def get_unit_economics_report(
    store_id: int | None = Query(default=None),
    query: str = Query(default=""),
    profitability: str = Query(default="all"),
    limit: int = Query(default=100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    owner_id = get_cabinet_owner_id(current_user)
    stores_result = await db.execute(select(Store).where(Store.user_id == owner_id))
    store_entities = stores_result.scalars().all()
    history_service = EconomicsHistoryService(db)
    today = datetime.now(timezone.utc).date()
    user_stores = []
    for store in store_entities:
        current_snapshot = await history_service.get_store_economics_for_date(store=store, as_of=today)
        user_stores.append(
            {
                "id": store.id,
                "name": store.name,
                "client_id": store.client_id,
                "api_key_encrypted": store.api_key_encrypted,
                "economics_vat_mode": current_snapshot.vat_mode,
                "economics_tax_mode": current_snapshot.tax_mode,
                "economics_tax_rate": current_snapshot.tax_rate,
                "economics_default_sale_price_gross": store.economics_default_sale_price_gross,
            }
        )
    user_store_ids = {int(store["id"]) for store in user_stores}

    if store_id is not None and store_id not in user_store_ids:
        raise HTTPException(status_code=404, detail="Магазин не найден")

    result = await DashboardUnitEconomicsService().build_report(
        db=db,
        user_id=owner_id,
        stores=user_stores,
        store_id=store_id,
        query=query,
        profitability=profitability,
        limit=limit,
        warehouse_mode=(await SettingsService(db).get_settings(current_user.id)).warehouse_mode,
    )
    return result
